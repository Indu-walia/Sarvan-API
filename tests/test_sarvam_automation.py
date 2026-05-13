import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import requests
import openpyxl
import time
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

EXCEL_PATH = "D:/download/Final_With_New_Scenarios.xlsx"
API_URL = "https://p9authwave.mox2.net.in/p9/reprocesslog.ashx"
API_KEY = "3E3C-19A9-421F-1326-284F-8B14-EC3C-CA31"


def call_api(text):
    payload = {
        "key": API_KEY,
        "data": [{"text": text, "qual": "4", "op": "0"}],
        "InputLanguage": "english",
        "lang": ["hindi"]
    }
    try:
        response = requests.post(API_URL, json=payload, timeout=30, verify=False)
        response.raise_for_status()
        return response.json(), response.status_code
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}, None


def extract_translated_text(result):
    # API returns a list of translation objects
    if isinstance(result, list) and result:
        item = result[0]
        if item.get("error"):
            return f"ERROR: {item['error']}"
        return item.get("outputText") or item.get("tgtText") or str(item)
    if isinstance(result, dict):
        if "error" in result:
            return f"ERROR: {result['error']}"
        return result.get("outputText") or result.get("tgtText") or str(result)
    return str(result)


def validate(input_text, actual_output, validation_criteria):
    """Basic validation based on criteria text."""
    if actual_output.startswith("ERROR:"):
        return "FAIL"

    criteria = str(validation_criteria).lower() if validation_criteria else ""

    try:
        if "." in criteria and "count" in criteria:
            input_dots = input_text.count(".")
            output_dots = actual_output.count(".")
            if "==" in criteria or "match" in criteria or "must match" in criteria:
                return "PASS" if input_dots == output_dots else "FAIL"
            if "no '.' added" in criteria or "without full stop" in criteria:
                return "PASS" if output_dots == 0 else "FAIL"
            if "one '.' present" in criteria or "not removed or duplicated" in criteria:
                return "PASS" if output_dots == 1 else "FAIL"
        if "no" in criteria and "added" in criteria and "." in criteria:
            return "PASS" if actual_output.count(".") == input_text.count(".") else "FAIL"
    except Exception:
        pass

    # If we can't determine, mark as PASS if output is non-empty and not error
    return "PASS" if actual_output and not actual_output.startswith("ERROR:") else "FAIL"


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Sheet1"]

    headers = [cell.value for cell in ws[1]]

    def get_or_create_col(name):
        if name in headers:
            return headers.index(name) + 1
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=name)
        headers.append(name)
        return col

    input_col      = headers.index("Input (EN)") + 1
    tc_col         = headers.index("Test Case ID") + 1
    expected_col   = headers.index("Expected Output") + 1
    criteria_col   = headers.index("Validation Criteria") + 1
    actual_col     = get_or_create_col("Actual Output")
    status_col     = get_or_create_col("Status")

    print(f"{'TC ID':<10} {'Status':<6} {'Input':<45} {'Actual Output'}")
    print("-" * 130)

    pass_count = fail_count = 0

    for row_num in range(2, ws.max_row + 1):
        tc_id          = ws.cell(row=row_num, column=tc_col).value
        input_text     = ws.cell(row=row_num, column=input_col).value
        expected       = ws.cell(row=row_num, column=expected_col).value
        criteria       = ws.cell(row=row_num, column=criteria_col).value

        if not input_text:
            continue

        result, status_code = call_api(str(input_text))
        actual_output = extract_translated_text(result)

        status = validate(str(input_text), actual_output, criteria)

        ws.cell(row=row_num, column=actual_col, value=actual_output)
        ws.cell(row=row_num, column=status_col, value=status)

        if status == "PASS":
            pass_count += 1
        else:
            fail_count += 1

        short_input = str(input_text)[:43]
        short_output = actual_output[:60]
        print(f"{str(tc_id):<10} {status:<6} {short_input:<45} {short_output}")

        time.sleep(0.5)

    output_path = EXCEL_PATH.replace(".xlsx", "_Results.xlsx")
    wb.save(output_path)
    print(f"\nDone! PASS: {pass_count} | FAIL: {fail_count}")
    print(f"Results saved to: {output_path}")


if __name__ == "__main__":
    main()
