import re
import sys
import html
import warnings
import requests
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font

warnings.filterwarnings('ignore')

API_URL = 'https://p9authwave.mox2.net.in/p9/reprocesslog.ashx'
API_KEY = '3642-E582-D3CC-49FA-FC38-F64C-0EC0-BCBB'

GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
BOLD  = Font(bold=True)

TAG_RE             = re.compile(r'<(/?[a-zA-Z]+\d*)\s*/?>', re.IGNORECASE)
BR_RE              = re.compile(r'^<br\s*/?>$', re.IGNORECASE)
TAG_PLACEHOLDER_RE = re.compile(r'^</?[a-zA-Z]+\d+>$')


def call_api(text):
    """Return the full list of response objects (one per sentence/br the API detects)."""
    payload = {
        'key': API_KEY,
        'data': [{'text': text, 'qual': '4', 'op': '0'}],
        'InputLanguage': 'english',
        'lang': ['hindi']
    }
    resp = requests.post(API_URL, json=payload, timeout=30, verify=False)
    resp.raise_for_status()
    return resp.json()


def is_br_object(d):
    """Return True if this API response object represents a <br> separator."""
    src = d.get('srcText', '')
    return bool(re.match(r'^\s*<br\s*/?>\s*$', src, re.IGNORECASE))


def extract_tags(text):
    names = TAG_RE.findall(text)
    return ','.join(f'<{n}>' for n in names if not re.match(r'^br\d*$', n, re.IGNORECASE))


def tag_name_list(tag_str):
    return TAG_RE.findall(str(tag_str) if tag_str else '')


def compare_tags(expected, actual):
    return 'Pass' if tag_name_list(expected) == tag_name_list(actual) else 'Fail'


def compute_output_values(tokens):
    return ','.join(
        t['srcText'] for t in tokens
        if not re.match(r'^<br\d*>$', t['context'], re.IGNORECASE)
    )


def unescape_expected(s):
    s = str(s)
    s = s.replace('\\\\"', '"')
    s = s.replace('\\"', '"')
    s = re.sub(r'\\+"', '"', s)
    return s.strip().strip('"').strip()


def normalize_val(s):
    """Normalize a value for comparison: decode HTML entities, unescape backslash-quotes, collapse whitespace."""
    s = str(s)
    s = html.unescape(s)            # &lt;b&gt; → <b>, &amp; → &, etc.
    s = s.replace('\\\\"', '"').replace('\\"', '"')
    s = re.sub(r'\\+"', '"', s)
    return re.sub(r'\s+', '', s)


def br_spacing_matches(source_text, br_objs, br_idx):
    """
    Compare the space BEFORE the nth <br> in source vs the nth <br> API response object.

    The API reports each <br> as a separate response object. Its srcText has a
    leading space only when the original source had a space immediately before that <br>.
    A space AFTER a <br> in source becomes the leading space of the NEXT br_obj,
    so we compare only space_before.

    Returns True  (Pass) if source space_before == API br_obj space_before.
    Returns False (Fail) if they differ.
    """
    src_matches = list(re.finditer(r'<br\s*/?>', source_text, re.IGNORECASE))
    if br_idx >= len(src_matches):
        return True

    m = src_matches[br_idx]
    src_space_before = m.start() > 0 and source_text[m.start() - 1] == ' '

    if br_idx >= len(br_objs):
        return True

    api_space_before = br_objs[br_idx].get('srcText', '').startswith(' ')

    return src_space_before == api_space_before


def _smart_comma_split(s):
    """
    Split string by comma, but re-merge adjacent HTML content tags that were split.
    Tag placeholders like <n1>, <an1> are kept separate (not merged).

    e.g. "<n1>,<n2>"                      → ["<n1>", "<n2>"]      (placeholders — keep separate)
    e.g. "<span class=\"x\">,</span>,9"   → ["<span class=\"x\"></span>", "9"]  (HTML — merge)
    e.g. "17,000"                          → ["17", "000"]          (plain values)
    """
    raw = [p.strip() for p in s.split(',') if p.strip()]
    result = []
    i = 0
    while i < len(raw):
        part = raw[i]
        while (i + 1 < len(raw)
               and part.endswith('>')
               and raw[i + 1].startswith('<')
               and not TAG_PLACEHOLDER_RE.match(part)
               and not TAG_PLACEHOLDER_RE.match(raw[i + 1])):
            i += 1
            part = part + raw[i]
        result.append(part)
        i += 1
    return result


def compare_output(expected, tokens, br_objs, source_text, suffix=''):
    """
    Validate InoutputText for non-pipe rows.

    br_objs : list of <br> response objects from the API (for spacing comparison).
    For each expected part:
      - <br> variant     → compare source spacing vs API br object spacing (must match)
      - tag placeholder  → compare against token context
      - HTML/plain value → compare against token srcText
    """
    if not expected and expected != 0:
        return 'Pass'

    # No tokens — API preserved the text as-is without tagging (e.g. #{...}# placeholders).
    # Verify the expected value appears in the original source text.
    non_br_check = [t for t in tokens
                    if not re.match(r'^<br\d*>$', t['context'], re.IGNORECASE)]
    if not non_br_check:
        return 'Pass' if str(expected).strip() in source_text else 'Fail'

    # Excel stores numbers like "17,000" as integer 17000 (no comma).
    # Compare by stripping formatting commas from both sides.
    if isinstance(expected, (int, float)):
        non_br = [t for t in tokens
                  if not re.match(r'^<br\d*>$', t['context'], re.IGNORECASE)]
        act_joined  = ','.join(t['srcText'] for t in non_br)
        exp_norm    = str(int(expected))
        act_norm    = act_joined.replace(',', '').replace(' ', '')
        return 'Pass' if exp_norm == act_norm else 'Fail'

    non_br_tokens = [t for t in tokens
                     if not re.match(r'^<br\d*>$', t['context'], re.IGNORECASE)]

    exp_str = unescape_expected(expected)
    sep     = '|' if '|' in exp_str else ','
    br_idx  = 0

    ctx_parts = [t['context'] for t in non_br_tokens]

    # Build actual srcText parts using smart split (preserves HTML but keeps placeholders separate)
    act_joined = ','.join(t['srcText'] for t in non_br_tokens)
    act_parts  = _smart_comma_split(act_joined)

    # Build expected parts using smart split then handle each by type
    if sep == ',':
        exp_parts = _smart_comma_split(exp_str)
    else:
        exp_parts = [p.strip() for p in exp_str.split('|')]

    tok_idx = 0

    for exp_val in exp_parts:
        clean = exp_val.strip().strip('"').strip()
        if not clean:
            continue

        if BR_RE.match(clean):
            if not br_spacing_matches(source_text, br_objs, br_idx):
                return 'Fail'
            br_idx += 1

        elif TAG_PLACEHOLDER_RE.match(clean):
            actual_ctx = ctx_parts[tok_idx] if tok_idx < len(ctx_parts) else ''
            tok_idx += 1
            if normalize_val(clean) != normalize_val(actual_ctx):
                return 'Fail'

        else:
            ne = normalize_val(clean)
            act_val = act_parts[tok_idx] if tok_idx < len(act_parts) else ''
            na = normalize_val(act_val)
            if ne == na:
                tok_idx += 1
                continue
            if suffix and ne == normalize_val(act_val + suffix):
                tok_idx += 1
                continue
            # Not a token match — check if it is contextual literal text
            # present in the original source (e.g. ₹ prefix before a variable).
            # If found in source, accept without consuming a token slot.
            if ne and ne in normalize_val(source_text):
                continue   # tok_idx NOT advanced — no token was consumed
            tok_idx += 1
            return 'Fail'

    return 'Pass'


def compare_output_pipe(expected, segment_tokens_list, br_objs, source_text):
    """
    Validate pipe-separated InoutputText.

    segment_tokens_list : token lists from sentence-type API response objects only.
    br_objs             : <br>-type API response objects (for spacing comparison).
    source_text         : original source text.

    '|' is the sentence-break separator — never causes a Fail by itself.
    Each '|'-separated part:
      - <br> part        → compare source spacing vs API br object spacing
      - tag placeholder  → compare against current sentence's token context
      - plain/HTML value → comma-separated sub-values vs current sentence's tokens
    Non-<br> parts advance to the next sentence; <br> parts do not.
    """
    if not expected:
        return 'Pass'

    exp_str   = unescape_expected(expected)
    exp_parts = [p.strip() for p in exp_str.split('|')]

    br_idx  = 0
    seg_idx = 0
    tok_idx = 0

    seg_non_br = [
        [t for t in tokens if not re.match(r'^<br\d*>$', t['context'], re.IGNORECASE)]
        for tokens in segment_tokens_list
    ]

    for exp_val in exp_parts:
        clean = exp_val.strip().strip('"').strip()
        if not clean:
            continue

        if BR_RE.match(clean):
            if not br_spacing_matches(source_text, br_objs, br_idx):
                return 'Fail'
            br_idx += 1

        elif TAG_PLACEHOLDER_RE.match(clean):
            cur = seg_non_br[seg_idx] if seg_idx < len(seg_non_br) else []
            ctx = cur[tok_idx]['context'] if tok_idx < len(cur) else ''
            if normalize_val(clean) != normalize_val(ctx):
                return 'Fail'
            seg_idx += 1
            tok_idx  = 0

        else:
            cur    = seg_non_br[seg_idx] if seg_idx < len(seg_non_br) else []
            merged = re.sub(r'>\s*,\s*<', '><', clean)
            raw    = [p.strip() for p in merged.split(',') if p.strip()]
            sub_parts = []
            i = 0
            while i < len(raw):
                part = raw[i]
                while (i + 1 < len(raw)
                       and part.endswith('>')
                       and raw[i + 1].startswith('<')
                       and not TAG_PLACEHOLDER_RE.match(part)
                       and not TAG_PLACEHOLDER_RE.match(raw[i + 1])):
                    i += 1
                    part = part + raw[i]
                sub_parts.append(part)
                i += 1

            for sub in sub_parts:
                act_tok = cur[tok_idx] if tok_idx < len(cur) else {}
                tok_idx += 1
                if TAG_PLACEHOLDER_RE.match(sub):
                    ctx = act_tok.get('context', '')
                    if normalize_val(sub) != normalize_val(ctx):
                        return 'Fail'
                else:
                    act = act_tok.get('srcText', '')
                    if normalize_val(sub) != normalize_val(act):
                        return 'Fail'
            seg_idx += 1
            tok_idx  = 0

    return 'Pass'


def check_not_contains(exp_not_contain, output_text):
    """
    Return 'Pass' / 'Fail' based on values in the Not-Contains column.

    Two kinds of values are handled:
      #{...}# patterns  → MUST be preserved in outputText (present = Pass, absent = Fail)
      everything else   → must NOT appear in outputText (present = Fail, absent = Pass)

    Empty / None → 'Pass' (skip check).
    """
    if not exp_not_contain:
        return 'Pass'
    exp_str = unescape_expected(exp_not_contain)
    if not exp_str:
        return 'Pass'
    values = [v.strip() for v in exp_str.split(',') if v.strip()]
    for val in values:
        if re.match(r'^#\{[^}]+\}#$', val):
            # Custom placeholder — must be preserved (present in outputText)
            if val not in output_text:
                return 'Fail'
        else:
            # Tag placeholder / text — must NOT appear in outputText
            if val in output_text:
                return 'Fail'
    return 'Pass'


def main():
    wb = openpyxl.load_workbook('TaggingAutomationData.xlsx')
    ws = wb.active

    new_headers = [
        'ActualEnglishText',
        'ActualIntgtText', 'IntgtText_Result',
        'ActualInsrcText', 'InsrcText_Result',
        'ActualInoutputText', 'InoutputText_Result',
        'NotContains_Result',
        'Overall_Result'
    ]
    start_col = ws.max_column + 1
    for idx, h in enumerate(new_headers):
        cell = ws.cell(1, start_col + idx, h)
        cell.font = BOLD

    for row in range(2, ws.max_row + 1):
        sno        = ws.cell(row, 1).value
        src_text   = ws.cell(row, 2).value or ''
        exp_intgt  = ws.cell(row, 3).value
        exp_insrc       = ws.cell(row, 4).value
        exp_output      = ws.cell(row, 5).value
        exp_not_contain = ws.cell(row, 6).value

        print(f'Processing row {row} (Sno={sno})...', end=' ', flush=True)

        try:
            all_data = call_api(src_text)
        except Exception as e:
            for idx in range(len(new_headers)):
                ws.cell(row, start_col + idx, f'ERROR: {e}')
            print('ERROR')
            continue

        # Separate sentence objects from <br> separator objects
        sentence_objs = [d for d in all_data if not is_br_object(d)]
        br_objs       = [d for d in all_data if is_br_object(d)]

        data         = sentence_objs[0] if sentence_objs else all_data[0]
        src_text_api = data.get('srcText', '')
        suffix       = data.get('suffix', '')

        # Combine tokens from ALL sentence objects — the API may split one long sentence
        # into multiple objects, and tagged values can appear in any of them.
        all_sentence_tokens = [t for d in sentence_objs for t in d.get('tokens', [])]

        exp_str_check = unescape_expected(exp_output) if exp_output else ''
        is_pipe_row   = '|' in exp_str_check

        if is_pipe_row:
            # Each sentence object → one pipe segment; br_objs used for spacing checks
            segment_tokens_list = [d.get('tokens', []) for d in sentence_objs]
            seg_tgt_tag_parts   = [extract_tags(d.get('tgtText', '')) for d in sentence_objs]
            seg_src_tag_parts   = [extract_tags(d.get('srcText', '')) for d in sentence_objs]

            actual_intgt  = ','.join(t for t in seg_tgt_tag_parts if t)
            actual_insrc  = ','.join(t for t in seg_src_tag_parts if t)
            actual_output = compute_output_values(all_sentence_tokens)

            intgt_result  = compare_tags(exp_intgt, actual_intgt)
            insrc_result  = compare_tags(exp_insrc, actual_insrc)
            output_result = compare_output_pipe(exp_output, segment_tokens_list,
                                                br_objs, src_text)
        else:
            # Non-pipe: use combined tags and tokens from all sentence objects
            all_tgt_tags  = ','.join(filter(None, [extract_tags(d.get('tgtText', '')) for d in sentence_objs]))
            all_src_tags  = ','.join(filter(None, [extract_tags(d.get('srcText', '')) for d in sentence_objs]))

            actual_intgt  = all_tgt_tags
            actual_insrc  = all_src_tags
            actual_output = compute_output_values(all_sentence_tokens)

            intgt_result  = compare_tags(exp_intgt, actual_intgt)
            insrc_result  = compare_tags(exp_insrc, actual_insrc)
            output_result = compare_output(exp_output, all_sentence_tokens, br_objs, src_text, suffix)

        # NotContains check against the final outputText.
        # #{...}# patterns → must BE preserved (present = Pass)
        # tag placeholders → must NOT appear (present = Fail)
        full_output_text = ' '.join(d.get('outputText', '') for d in sentence_objs)
        not_contains_result = check_not_contains(exp_not_contain, full_output_text)

        overall = 'Pass' if (
            intgt_result == insrc_result == output_result == not_contains_result == 'Pass'
        ) else 'Fail'

        values = [
            src_text_api,
            actual_intgt,  intgt_result,
            actual_insrc,  insrc_result,
            actual_output, output_result,
            not_contains_result,
            overall
        ]
        for idx, val in enumerate(values):
            cell = ws.cell(row, start_col + idx, val)
            if val in ('Pass', 'Fail'):
                cell.fill = GREEN if val == 'Pass' else RED

        print(f'IntgtText={intgt_result}, InsrcText={insrc_result}, Output={output_result} → {overall}')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_file = f'TaggingAutomationData_Results_{timestamp}.xlsx'
    wb.save(out_file)
    print(f'\nSaved: {out_file}')


if __name__ == '__main__':
    sys.stdout.reconfigure(encoding='utf-8')
    main()
