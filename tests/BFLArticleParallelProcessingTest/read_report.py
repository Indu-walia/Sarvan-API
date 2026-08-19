from openpyxl import load_workbook
from pathlib import Path

wb_path = Path('report.xlsx')
if not wb_path.exists():
    print(f"Report not found: {wb_path.resolve()}")
    raise SystemExit(1)

wb = load_workbook(wb_path)
s = wb.active
rows = list(s.iter_rows(values_only=True))
if not rows:
    print('Empty report')
    raise SystemExit(0)

headers = rows[0]
print('Headers:', headers)

failures = []
success = 0
for r in rows[1:]:
    rec = dict(zip(headers, r))
    code = rec.get('status_code')
    try:
        ok = code is not None and int(code) >= 200 and int(code) < 300
    except Exception:
        ok = False
    if ok:
        success += 1
    else:
        failures.append(rec)

print(f"Total rows: {len(rows)-1}, Success: {success}, Failures: {len(failures)}")
if failures:
    print('\nFailures (first 20):')
    for i, f in enumerate(failures[:20], 1):
        print(f"{i}. file_name={f.get('file_name')}, status_code={f.get('status_code')}, error={f.get('error')}")
        snippet = (f.get('response_snippet') or '').strip()
        if snippet:
            print('   response_snippet:', snippet[:1000])

print('\nDone')
