#!/usr/bin/env python3
"""Calculate article-wise turnaround time from Process9 billing grid reports
(the "ArticleId, Status, Language, RequestType, RequestTime, ResponseTime,
Quality, WordCount, Department, Error, ClientName" export format) and write
an Excel report with one row per article submission.

Usage:
    python scripts/calculate_article_time.py "Article_Report.xlsx" "Article_Report (1).xlsx"
    python scripts/calculate_article_time.py "Article_Report.xlsx" "Article_Report (1).xlsx" --output "Article_Wise_Time_Report.xlsx"
"""
import argparse
import re
import statistics
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TIME_FORMAT = "%d-%m-%Y %H:%M:%S"


def load(path):
    wb = load_workbook(path)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[4]
    data = []
    for r in rows[5:]:
        if r[0] is None:
            continue
        data.append(dict(zip(header, r)))
    return data


def parse(ts):
    return datetime.strptime(ts, TIME_FORMAT)


def build_records(all_data):
    delivered_only = [d for d in all_data if d["Status"] == "Delivered"]

    by_article = defaultdict(list)
    for d in delivered_only:
        by_article[d["ArticleId"]].append(d)

    records = []
    for article_id, rows in by_article.items():
        m = re.match(r"(.+?)_loop(\d+)_\d+$", article_id)
        base, loop = (m.group(1), m.group(2)) if m else (article_id, "")

        client = rows[0]["ClientName"]
        req_time = min(parse(r["RequestTime"]) for r in rows)

        lang_durations, resp_times = [], []
        for r in rows:
            rt, rs = parse(r["RequestTime"]), parse(r["ResponseTime"])
            lang_durations.append((rs - rt).total_seconds() / 60)
            resp_times.append(rs)

        first_resp = min(resp_times)
        last_resp = max(resp_times)
        total_duration_min = (last_resp - req_time).total_seconds() / 60

        records.append({
            "ArticleId": article_id,
            "BaseArticle": base,
            "Loop": loop,
            "Client": client,
            "RequestTime": req_time,
            "FirstResponseTime": first_resp,
            "LastResponseTime": last_resp,
            "TotalDurationMin": round(total_duration_min, 2),
            "FastestLangMin": round(min(lang_durations), 2),
            "SlowestLangMin": round(max(lang_durations), 2),
            "AvgLangMin": round(statistics.mean(lang_durations), 2),
            "LanguagesDelivered": len(rows),
            "WordCount": int(rows[0]["WordCount"]) if rows[0]["WordCount"] else 0,
        })

    records.sort(key=lambda r: (r["Client"], r["BaseArticle"], r["Loop"]))
    return records


def write_excel(records, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Article-wise Time"

    headers = [
        "ArticleId", "Base Article", "Loop", "Client", "Request Time",
        "First Response", "Last Response", "Total Duration (min)",
        "Fastest Language (min)", "Slowest Language (min)", "Avg Language (min)",
        "Languages Delivered", "Word Count",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0E7C7B", end_color="0E7C7B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    dt_fmt = "dd-mm-yyyy hh:mm:ss"

    for r in records:
        row = [
            r["ArticleId"], r["BaseArticle"], r["Loop"], r["Client"],
            r["RequestTime"], r["FirstResponseTime"], r["LastResponseTime"],
            r["TotalDurationMin"], r["FastestLangMin"], r["SlowestLangMin"], r["AvgLangMin"],
            r["LanguagesDelivered"], r["WordCount"],
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col in (5, 6, 7):
            ws.cell(row=row_idx, column=col).number_format = dt_fmt

    for idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))] + [len(str(ws.cell(row=i, column=idx).value)) for i in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 3, 26)

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary_headers = ["Client", "Articles", "Avg Total Duration (min)", "Min (min)", "Max (min)", "Total Words"]
    summary.append(summary_headers)
    for col_idx in range(1, len(summary_headers) + 1):
        cell = summary.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    by_client = defaultdict(list)
    for r in records:
        by_client[r["Client"]].append(r)

    for client, recs in sorted(by_client.items()):
        durs = [r["TotalDurationMin"] for r in recs if r["TotalDurationMin"] is not None]
        words = sum(r["WordCount"] for r in recs)
        summary.append([
            client, len(recs),
            round(statistics.mean(durs), 2) if durs else None,
            round(min(durs), 2) if durs else None,
            round(max(durs), 2) if durs else None,
            words,
        ])

    all_durs = [r["TotalDurationMin"] for r in records if r["TotalDurationMin"] is not None]
    if all_durs:
        all_words = sum(r["WordCount"] for r in records)
        summary.append([
            "Combined", len(records),
            round(statistics.mean(all_durs), 2), round(min(all_durs), 2), round(max(all_durs), 2), all_words,
        ])
        summary.cell(row=summary.max_row, column=1).font = Font(bold=True)

    for idx, header in enumerate(summary_headers, start=1):
        max_len = max(
            [len(str(header))] + [len(str(summary.cell(row=i, column=idx).value)) for i in range(2, summary.max_row + 1)]
        )
        summary.column_dimensions[get_column_letter(idx)].width = min(max_len + 3, 26)

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("reports", nargs="+", help="One or more Process9 billing grid report .xlsx files")
    parser.add_argument("--output", default="Article_Wise_Time_Report.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    all_data = []
    for path in args.reports:
        all_data.extend(load(path))

    records = build_records(all_data)
    write_excel(records, args.output)

    print(f"Saved: {args.output}")
    print(f"Article rows: {len(records)}")


if __name__ == "__main__":
    main()
