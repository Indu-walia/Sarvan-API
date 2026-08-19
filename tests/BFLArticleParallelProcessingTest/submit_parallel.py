import argparse
import csv
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def find_json_files(base_dir: Path, recursive: bool) -> List[Path]:
    if not base_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {base_dir}")
    if not base_dir.is_dir():
        raise NotADirectoryError(f"Input path is not a directory: {base_dir}")

    if recursive:
        candidates = [path for path in base_dir.rglob("*") if path.is_file()]
    else:
        candidates = [path for path in base_dir.iterdir() if path.is_file()]

    return candidates


def load_payload(file_path: Path) -> Optional[Dict[str, Any]]:
    try:
        with file_path.open("r", encoding="utf-8") as stream:
            payload = json.load(stream)
            if not isinstance(payload, dict):
                return None
            return payload
    except Exception:
        return None


def send_payload(session: requests.Session, url: str, headers: Dict[str, str], payload: Dict[str, Any]) -> Dict[str, Any]:
    start = time.monotonic()
    try:
        response = session.post(url, headers=headers, json=payload, timeout=60)
        elapsed = (time.monotonic() - start) * 1000
        return {
            "status_code": response.status_code,
            "reason": response.reason,
            "response_text": response.text,
            "elapsed_ms": round(elapsed, 1),
            "error": None,
        }
    except Exception as exc:
        elapsed = (time.monotonic() - start) * 1000
        return {
            "status_code": None,
            "reason": None,
            "response_text": "",
            "elapsed_ms": round(elapsed, 1),
            "error": str(exc),
        }


def write_report(report_path: Path, rows: List[Dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Load Report"

    headers = [
        "file_name",
        "article_id",
        "article_title",
        "status_code",
        "reason",
        "elapsed_ms",
        "error",
        "response_snippet",
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append([
            row.get("file_name", ""),
            row.get("article_id", ""),
            row.get("article_title", ""),
            row.get("status_code", ""),
            row.get("reason", ""),
            row.get("elapsed_ms", ""),
            row.get("error", ""),
            row.get("response_snippet", ""),
        ])

    for idx, column_cells in enumerate(sheet.columns, start=1):
        max_length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=10,
        )
        width = min(max_length + 4, 70)
        sheet.column_dimensions[get_column_letter(idx)].width = width

    workbook.save(report_path)


def write_csv_report(report_path: Path, rows: List[Dict[str, Any]]) -> None:
    headers = [
        "file_name",
        "article_id",
        "article_title",
        "status_code",
        "reason",
        "elapsed_ms",
        "error",
        "response_snippet",
    ]
    with report_path.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "file_name": row.get("file_name", ""),
                    "article_id": row.get("article_id", ""),
                    "article_title": row.get("article_title", ""),
                    "status_code": row.get("status_code", ""),
                    "reason": row.get("reason", ""),
                    "elapsed_ms": row.get("elapsed_ms", ""),
                    "error": row.get("error", ""),
                    "response_snippet": row.get("response_snippet", ""),
                }
            )


def clear_es_indices(es_base: str, keys: List[str], session: requests.Session) -> None:
    """Attempt to delete indices matching the given keys. Tries a few candidate names per key."""
    es_base = es_base.rstrip("/")
    for key in keys:
        candidates = [key, key.lower(), key.upper(), f"articles-{key}", f"articles-{key.lower()}"]
        deleted_any = False
        for cand in candidates:
            url = f"{es_base}/{cand}"
            try:
                r = session.delete(url, timeout=20)
                if r.status_code in (200, 202):
                    print(f"Elasticsearch: deleted index '{cand}' (status={r.status_code})")
                    deleted_any = True
                    break
                elif r.status_code == 404:
                    # not found; continue trying other candidate names
                    print(f"Elasticsearch: index '{cand}' not found (404)")
                else:
                    print(f"Elasticsearch: delete '{cand}' returned {r.status_code}: {r.text[:200]}")
            except Exception as exc:
                print(f"Elasticsearch: error deleting '{cand}': {exc}")
        if not deleted_any:
            print(f"Elasticsearch: no index deleted for key '{key}' (checked candidates)")


def delete_docs_by_key(es_base: str, key: str, session: requests.Session) -> None:
    """Delete documents matching the key using _delete_by_query across all indices.

    This uses a query_string search for the literal key to match any field containing it.
    """
    es_base = es_base.rstrip("/")
    # First fetch index list and run delete-by-query per index to avoid cluster-level endpoint issues
    cat_indices = f"{es_base}/_cat/indices?format=json"
    try:
        ir = session.get(cat_indices, timeout=30)
        if ir.status_code != 200:
            print(f"Elasticsearch: failed to list indices ({ir.status_code}): {ir.text[:200]}")
            return
        try:
            idx_list = [item.get("index") for item in ir.json() if item.get("index")]
        except Exception:
            print("Elasticsearch: unexpected response when listing indices")
            return
    except Exception as exc:
        print(f"Elasticsearch: error listing indices: {exc}")
        return

    payload = {
        "query": {
            "query_string": {
                "query": f'"{key}"'
            }
        }
    }

    deleted_total = 0
    for index in idx_list:
        url = f"{es_base}/{index}/_delete_by_query"
        try:
            r = session.post(url, json=payload, timeout=120)
            if r.status_code in (200, 201):
                try:
                    data = r.json()
                    deleted = data.get("deleted", data.get("total", None))
                except Exception:
                    deleted = None
                print(f"Elasticsearch: deleted from index '{index}' for key '{key}': deleted={deleted}")
                try:
                    if isinstance(deleted, int):
                        deleted_total += int(deleted)
                except Exception:
                    pass
            elif r.status_code == 404:
                # index not found or endpoint not available on this index
                continue
            else:
                print(f"Elasticsearch: _delete_by_query on index '{index}' returned {r.status_code}: {r.text[:200]}")
        except Exception as exc:
            print(f"Elasticsearch: error on index '{index}': {exc}")

    print(f"Elasticsearch: total deleted matching '{key}': {deleted_total}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Load all JSON files from a folder and submit them in parallel to the article API.")
    parser.add_argument(
        "--input-dir",
        default="Test data parallel processing",
        help="Folder containing JSON payload files (plain .json or any file with JSON content).",
    )
    parser.add_argument("--url", default="https://qa_article.mox2.net.in/Article", help="API endpoint URL")
    parser.add_argument("--api-key", default="7895-620F-FB0B-8853-0015-17C6-8B13-720A", help="X-API-KEY header value")
    parser.add_argument("--workers", type=int, default=10, help="Number of concurrent worker threads")
    parser.add_argument("--recursive", action="store_true", help="Recurse into subdirectories to find JSON payload files")
    parser.add_argument("--output-prefix", default="article_load_report", help="Output filename prefix for Excel report")
    parser.add_argument("--output-file", default=None, help="Excel report output file")
    parser.add_argument("--csv-file", default=None, help="Optional CSV report output file (omit to skip CSV output)")
    parser.add_argument("--dry-run", action="store_true", help="Validate JSON files without sending requests")
    parser.add_argument("--clear-es", action="store_true", help="Clear Elasticsearch indices for the provided keys before submitting")
    parser.add_argument("--no-pre-clear", action="store_true", help="Disable ES index + document cleanup before submitting (opt-out)")
    parser.add_argument("--filter-dept", default=None, help="Only submit payloads whose `departmentName` equals this value")
    parser.add_argument("--sequential", action="store_true", help="Send requests one-by-one instead of parallel workers")
    parser.add_argument("--filter-name", default=None, help="Only submit files whose filename contains this substring (case-insensitive)")
    parser.add_argument("--es-url", default="http://20.192.26.81:9200", help="Elasticsearch base URL")
    parser.add_argument(
        "--es-keys",
        default="929E-4237-1BE7-D479-F5D5-A39D-3B71-553C,7895-620F-FB0B-8853-0015-17C6-8B13-720A",
        help="Comma-separated list of ES index names or keys to clear",
    )
    parser.add_argument("--delete-docs", action="store_true", help="Run _delete_by_query for each key to remove matching documents")
    parser.add_argument("--delete-only", action="store_true", help="Run ES deletions only and exit (no submissions)")
    args = parser.parse_args()

    if args.output_file is None:
        args.output_file = f"{args.output_prefix}.xlsx"

    es_keys = [k.strip() for k in args.es_keys.split(",") if k.strip()]

    base_dir = Path(args.input_dir)
    files = find_json_files(base_dir, args.recursive)

    if not files:
        raise SystemExit(f"No files found in {base_dir}")

    print(f"Found {len(files)} candidate files in {base_dir}")

    records: List[Dict[str, Any]] = []
    json_files: List[Path] = []
    for file_path in files:
        payload = load_payload(file_path)
        if payload is None:
            records.append(
                {
                    "file_name": str(file_path.relative_to(base_dir)),
                    "article_id": "",
                    "article_title": "",
                    "status_code": None,
                    "reason": None,
                    "elapsed_ms": 0,
                    "error": "Invalid JSON or failed to parse",
                    "response_snippet": "",
                }
            )
            continue
        # If filename filter provided, skip files that don't match
        if args.filter_name:
            if args.filter_name.lower() not in file_path.name.lower():
                continue
        json_files.append(file_path)

    if not json_files:
        raise SystemExit("No valid JSON payload files found.")

    print(f"Will submit {len(json_files)} valid JSON files")

    if args.dry_run:
        for path in json_files:
            print(f"VALID JSON: {path.relative_to(base_dir)}")
        print("Dry run finished. No requests were sent.")
        write_report(Path(args.output_file), records)
        print(f"Wrote validation report to {args.output_file}")
        if args.csv_file:
            write_csv_report(Path(args.csv_file), records)
            print(f"Wrote CSV validation report to {args.csv_file}")
        return

    headers = {"X-API-KEY": args.api_key, "Content-Type": "application/json"}
    session = requests.Session()

    # By default run a pre-clear step unless user explicitly opts out with --no-pre-clear
    pre_clear_active = not getattr(args, "no_pre_clear", False)
    if pre_clear_active:
        print("Pre-clear enabled: running ES index deletion and document deletion for configured keys")
        print(f"Clearing Elasticsearch indices: {es_keys} at {args.es_url}")
        clear_es_indices(args.es_url, es_keys, session)
        if args.delete_docs:
            print(f"Deleting documents matching keys: {es_keys} at {args.es_url}")
            for k in es_keys:
                delete_docs_by_key(args.es_url, k, session)
        if args.delete_only:
            print("Delete-only requested; exiting after ES deletions.")
            return
    else:
        if args.clear_es:
            print(f"Clearing Elasticsearch indices: {es_keys} at {args.es_url}")
            clear_es_indices(args.es_url, es_keys, session)
        if args.delete_docs:
            print(f"Deleting documents matching keys: {es_keys} at {args.es_url}")
            for k in es_keys:
                delete_docs_by_key(args.es_url, k, session)
            if args.delete_only:
                print("Delete-only requested; exiting after ES deletions.")
                return

    file_to_payload = {}
    for file_path in json_files:
        payload = load_payload(file_path)
        if payload is None:
            continue
        # Optionally filter by departmentName key
        if args.filter_dept:
            dept = str(payload.get("departmentName", ""))
            if dept != args.filter_dept:
                # skip files that don't match the filter
                continue
        file_to_payload[file_path] = payload

    # Send requests either sequentially or in parallel
    if args.sequential:
        for file_path, payload in file_to_payload.items():
            result = send_payload(session, args.url, headers, payload)
            response_snippet = result["response_text"][:500].replace("\n", " ")
            record = {
                "file_name": str(file_path.relative_to(base_dir)),
                "article_id": str(payload.get("articleId", "")),
                "article_title": str(payload.get("articleTitle", "")),
                "status_code": result["status_code"],
                "reason": result["reason"],
                "elapsed_ms": result["elapsed_ms"],
                "error": result["error"],
                "response_snippet": response_snippet,
            }
            records.append(record)
            status = result["status_code"] or "ERROR"
            print(f"{record['file_name']}: {status} {record['elapsed_ms']}ms", f"error={record['error']}" if record["error"] else "")
    else:
        futures = {}
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            for file_path, payload in file_to_payload.items():
                futures[executor.submit(send_payload, session, args.url, headers, payload)] = (file_path, payload)

            for future in as_completed(futures):
                file_path, payload = futures[future]
                result = future.result()
                response_snippet = result["response_text"][:500].replace("\n", " ")
                record = {
                    "file_name": str(file_path.relative_to(base_dir)),
                    "article_id": str(payload.get("articleId", "")),
                    "article_title": str(payload.get("articleTitle", "")),
                    "status_code": result["status_code"],
                    "reason": result["reason"],
                    "elapsed_ms": result["elapsed_ms"],
                    "error": result["error"],
                    "response_snippet": response_snippet,
                }
                records.append(record)
                status = result["status_code"] or "ERROR"
                print(
                    f"{record['file_name']}: {status} {record['elapsed_ms']}ms",
                    f"error={record['error']}" if record["error"] else "",
                )

    write_report(Path(args.output_file), records)
    print(f"Saved Excel report to {args.output_file}")
    if args.csv_file:
        write_csv_report(Path(args.csv_file), records)
        print(f"Saved CSV report to {args.csv_file}")
    success_count = sum(1 for r in records if r["status_code"] and 200 <= int(r["status_code"]) < 300)
    failure_count = len(records) - success_count
    print(f"Summary: {success_count} success, {failure_count} failure over {len(records)} articles")


if __name__ == "__main__":
    main()
