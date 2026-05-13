import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import re
import requests
import openpyxl
import time

INPUT_EXCEL  = "D:/download/Final_With_New_Scenarios.xlsx"
OUTPUT_EXCEL = "D:/download/Sarvam_Manipuri_Results.xlsx"

API_URL     = "https://api.sarvam.ai/translate"
API_KEY     = "sk_6q9mn91f_NDW17uFDqDE5C5JLDtpR9iyL"
SRC_LANG    = "en-IN"
TGT_LANG    = "mni-IN"
MODEL       = "sarvam-translate:v1"

NEW_CASES = [
    # Punctuation variety
    ('TC_ADD_01', 'Question mark preserved',        'Are you ready?',                                          'Translated with one ?',                        "? not removed or duplicated"),
    ('TC_ADD_02', 'Exclamation preserved',          'Great job!',                                              'Translated with one !',                        'Exclamation mark intact'),
    ('TC_ADD_03', 'Comma preservation',             'Hello, how are you?',                                     'Translated with comma and ?',                  'Comma and ? must match input'),
    ('TC_ADD_04', 'Ellipsis preservation',          'Wait... let me check',                                    'Translated with ellipsis ...',                 'Three dots must remain unchanged'),
    ('TC_ADD_05', 'Semicolon preservation',         'Check one; check two',                                    'Translated with semicolon ;',                  'Semicolon not removed'),
    ('TC_ADD_06', 'Colon preservation',             'Note: This is important',                                 'Translated with colon :',                      'Colon not removed'),
    ('TC_ADD_07', 'Mixed punctuation',              'Hello! How are you? Fine.',                               'Translated with ! ? and . preserved',          '! ? . each count must match input'),
    ('TC_ADD_08', 'Parentheses preservation',       'Result (correct) has been verified',                      'Translated with ( ) intact',                   'Round brackets not removed'),
    ('TC_ADD_09', 'Square brackets',               '[Important] Please review this',                           'Translated with [ ] intact',                   'Square brackets preserved'),
    # Number formats
    ('TC_ADD_10', 'Decimal number',                 'Price is 99.99',                                          'Decimal number unchanged',                     'Numeric value 99.99 preserved exactly'),
    ('TC_ADD_11', 'Phone number',                   'Call at 9876543210',                                      '10-digit phone number unchanged',              'Phone number not translated or broken'),
    ('TC_ADD_12', 'Date format DD/MM/YYYY',         'Meeting on 25/04/2024',                                   'Date format unchanged',                        'Date string 25/04/2024 preserved'),
    ('TC_ADD_13', 'Time format AM/PM',              'Meeting at 10:30 AM',                                     'Time format unchanged',                        'Time and AM preserved'),
    ('TC_ADD_14', 'Percentage symbol',              'Score is 95%',                                            'Percentage intact',                            '% symbol and number unchanged'),
    ('TC_ADD_15', 'Negative number',                'Temperature is -5 degrees',                               'Negative sign preserved',                      'Minus sign and number unchanged'),
    # Special symbols
    ('TC_ADD_16', 'Hashtag handling',               '#TeamWork is the key',                                    'Hashtag label preserved',                      '#TeamWork not translated or broken'),
    ('TC_ADD_17', 'At-mention handling',            '@John please review this',                                '@John preserved as-is',                        '@John not altered'),
    ('TC_ADD_18', 'Dollar currency',                'Price is $500',                                           '$ symbol and number unchanged',                '$ and 500 intact'),
    ('TC_ADD_19', 'Euro currency text',             'Price is 200 euros',                                      'Currency context preserved',                   'Numeric value unchanged'),
    ('TC_ADD_20', 'Asterisk emphasis',              'Note: *important* point',                                 '* symbols preserved',                          'Asterisks not removed'),
    ('TC_ADD_21', 'Forward slash',                  'A/B testing is done',                                     'Slash preserved',                              '/ not removed'),
    ('TC_ADD_22', 'Underscore in identifier',       'file_name is ready',                                      'Underscore preserved',                         'file_name not broken'),
    ('TC_ADD_23', 'Ampersand symbol',               'Terms & Conditions apply',                                '& preserved or spelled out correctly',         '& symbol intact'),
    # Edge cases
    ('TC_ADD_24', 'Single word input',              'Hello',                                                   'Single word translated',                       'Non-empty output, no extra punctuation'),
    ('TC_ADD_25', 'All caps input',                 'THIS IS IMPORTANT',                                       'Translated meaningfully',                      'Non-empty meaningful output'),
    ('TC_ADD_26', 'Very long sentence',             'The development team has successfully completed the implementation of the new authentication module and all test cases are now passing.',
                                                                                                                'Full translation without truncation',          'Output non-empty and complete'),
    ('TC_ADD_27', 'Repeated words',                 'Very very very good',                                     'Repetition conveyed',                          'Non-empty output with context preserved'),
    ('TC_ADD_28', 'Only numbers input',             '12345',                                                   'Numbers unchanged or numeral form',            'Numeric meaning preserved'),
    # Code-like content
    ('TC_ADD_29', 'Error code in text',             'Error 404: Page not found',                               '404 preserved, rest translated',               'Numeric code 404 unchanged'),
    ('TC_ADD_30', 'Version number in text',         'Version 2.0.1 has been released',                         'Version number unchanged',                     '2.0.1 intact in output'),
    ('TC_ADD_31', 'Order ID in text',               'Order #ORD-2024-001 is confirmed',                        'Order ID preserved',                           '#ORD-2024-001 unchanged'),
    # More Hinglish
    ('TC_ADD_32', 'Hinglish with question mark',    'Kya yaar this is not working?',                           'Meaning preserved with ?',                     '? preserved, context intact'),
    ('TC_ADD_33', 'Hinglish office context',        'Kal meeting hai office mein',                             'Hindi-English mix translated meaningfully',    'Meaningful Hindi output'),
    ('TC_ADD_34', 'Hinglish request',               'Please dekho this file',                                  'Context preserved in translation',             'Non-empty output'),
    # Punctuation edge cases
    ('TC_ADD_35', 'Exclamation at end',             'Great work done!',                                        '! preserved at end',                           '! intact'),
    ('TC_ADD_36', 'Full stop and exclamation',      'Work is done. Well done!',                                '. and ! both preserved',                       '. and ! count must match input'),
    ('TC_ADD_37', 'No punctuation long',            'This is a simple sentence without any punctuation',       'Translated without adding punctuation',        'No . or ! added to output'),
    # HTML + combos
    ('TC_ADD_38', 'HTML with number',               '<p>Total items: 50</p>',                                  'HTML tags and number preserved',               'Tags intact, 50 unchanged'),
    ('TC_ADD_39', 'Template with OTP',              'Hello {{name}}, your OTP is 1234.',                       'Template and number preserved',                '{{name}} and 1234 unchanged'),
    ('TC_ADD_40', 'URL with query params',          'Visit https://site.com?id=1&page=2',                      'Full URL with params preserved',               'Complete URL unchanged'),
    # Whitespace
    ('TC_ADD_41', 'Leading spaces',                 '  Hello team',                                            'Leading space handled gracefully',             'Output non-empty and meaningful'),
    ('TC_ADD_42', 'Trailing spaces',                'Hello team  ',                                            'Trailing space handled gracefully',            'Output non-empty and meaningful'),
    # Formal/polite tone
    ('TC_ADD_43', 'Formal request',                 'Kindly share the updated document at the earliest.',      'Polite tone preserved',                        'Non-empty formal Hindi output'),
    ('TC_ADD_44', 'Question with please',           'Can you please check this issue?',                        'Polite ? preserved',                           '? intact, non-empty output'),
    # Number variety
    ('TC_ADD_45', 'Ordinal numbers',                '1st and 2nd tasks are done',                              'Ordinals preserved',                           '1st and 2nd unchanged'),
    ('TC_ADD_46', 'Spelled-out number',             'There are twenty five items',                             'Number context translated',                    'Non-empty output'),
    ('TC_ADD_47', 'Mixed numeral types',            'Call 1800-123-456 or visit Room 101',                     'Phone and room number preserved',              'Numerics unchanged'),
    # Currency combos
    ('TC_ADD_48', 'Multiple currencies',            'Pay $100 or 8000 rupees',                                 '$ and rupee value preserved',                  'Both numeric values unchanged'),
]


# ── API call ──────────────────────────────────────────────────────────────────

def call_api(text):
    try:
        r = requests.post(
            API_URL,
            headers={"Content-Type": "application/json", "api-subscription-key": API_KEY},
            json={"input": text, "source_language_code": SRC_LANG,
                  "target_language_code": TGT_LANG, "model": MODEL},
            timeout=30
        )
        r.raise_for_status()
        return r.json()
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}


def extract_output(result):
    if isinstance(result, dict):
        if "translated_text" in result:
            return result["translated_text"]
        if "error" in result:
            return f"ERROR: {result['error']}"
    return str(result)


# ── Validation (uses BOTH Expected Output + Validation Criteria) ──────────────

def validate(inp, actual, expected_output, criteria):
    if not actual or actual.startswith("ERROR:"):
        return "FAIL"
    if not actual.strip():
        return "FAIL"

    exp  = str(expected_output).lower() if expected_output else ""
    crit = str(criteria).lower()        if criteria        else ""

    # Hindi danda '।' (U+0964) is the Hindi equivalent of '.'
    # Count both when validating full stops
    def dot_count(text):
        return text.count(".") + text.count("।")

    # ── Full stop rules ────────────────────────────────────────────────────
    # "without full stop" / "no . added" → output must NOT gain extra full stops
    if "without full stop" in exp or ("no" in crit and "added" in crit and "." in crit):
        return "PASS" if dot_count(actual) == dot_count(inp) else "FAIL"

    # "one full stop" / "one '.'" → exactly 1 full stop (. or ।) in output
    if "one full stop" in exp or "one '.'" in crit:
        return "PASS" if dot_count(actual) == 1 else "FAIL"

    # ". and ! both preserved" or "each count must match" → strict count match
    if ("and !" in exp and "." in exp and "both preserved" in exp) or "each count must match" in crit:
        dots_ok  = dot_count(actual) == dot_count(inp)
        excl_ok  = actual.count("!") == inp.count("!")
        quest_ok = actual.count("?") == inp.count("?")
        return "PASS" if (dots_ok and excl_ok and quest_ok) else "FAIL"

    # "with full stop" or "full stop" → at least one full stop (. or ।) in output
    if "full stop" in exp or "full stop" in crit:
        return "PASS" if dot_count(actual) > 0 else "FAIL"

    # ── Specific punctuation ───────────────────────────────────────────────
    if "?" in exp or "?" in crit:
        return "PASS" if actual.count("?") == inp.count("?") else "FAIL"

    if "!" in exp or "exclamation" in crit:
        return "PASS" if actual.count("!") == inp.count("!") else "FAIL"

    if ";" in exp or "semicolon" in crit:
        return "PASS" if ";" in actual else "FAIL"

    if "colon" in crit and "not removed" in crit:
        return "PASS" if ":" in actual else "FAIL"

    if "( )" in exp or "round brackets" in crit:
        return "PASS" if ("(" in actual and ")" in actual) else "FAIL"

    if "[ ]" in exp or "square brackets" in crit:
        return "PASS" if ("[" in actual and "]" in actual) else "FAIL"

    if "..." in exp or "ellipsis" in crit:
        return "PASS" if "..." in actual else "FAIL"

    if "%" in exp or "%" in crit:
        return "PASS" if "%" in actual else "FAIL"

    # ── Token preservation ─────────────────────────────────────────────────
    tokens_to_check = []

    # URLs
    tokens_to_check += re.findall(r'https?://\S+', inp)

    # Emails
    tokens_to_check += re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', inp)

    # Template variables {{...}}
    tokens_to_check += re.findall(r'\{\{[^}]+\}\}', inp)

    # IDs like #ORD-2024-001
    tokens_to_check += re.findall(r'#[A-Z][A-Z0-9\-]+', inp)

    # When expected/criteria says "unchanged" / "preserved" / "intact"
    if any(w in exp or w in crit for w in ("unchanged", "preserved", "intact", "not broken", "not altered")):
        # phone numbers (7+ digits)
        tokens_to_check += re.findall(r'\b\d{7,}\b', inp)
        # dates  dd/mm/yyyy
        tokens_to_check += re.findall(r'\b\d{1,2}/\d{1,2}/\d{4}\b', inp)
        # version  x.y.z
        tokens_to_check += re.findall(r'\b\d+\.\d+\.\d+\b', inp)
        # decimal  99.99
        tokens_to_check += re.findall(r'\b\d+\.\d+\b', inp)
        # 3-4 digit codes (404, 101, 1234…)
        tokens_to_check += re.findall(r'\b\d{3,4}\b', inp)
        # ordinals  1st 2nd 3rd
        tokens_to_check += re.findall(r'\b\d+(?:st|nd|rd|th)\b', inp)
        # @mentions
        tokens_to_check += re.findall(r'@\w+', inp)

    if tokens_to_check:
        return "PASS" if all(t in actual for t in tokens_to_check) else "FAIL"

    # ── Emoji preservation ─────────────────────────────────────────────────
    if "emoji" in exp or "emoji" in crit:
        inp_emojis = [c for c in inp  if ord(c) > 0x1F300]
        out_emojis = [c for c in actual if ord(c) > 0x1F300]
        return "PASS" if inp_emojis == out_emojis else "FAIL"

    # ── "no punctuation added" ────────────────────────────────────────────
    if "without adding punctuation" in exp or ("no" in crit and "added" in crit):
        return "PASS" if dot_count(actual) == dot_count(inp) else "FAIL"

    # ── Default: non-empty output = PASS ──────────────────────────────────
    return "PASS" if actual.strip() else "FAIL"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws = wb["Sheet1"]

    headers = [cell.value for cell in ws[1]]

    def get_or_create_col(name):
        if name in headers:
            return headers.index(name) + 1
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=name)
        headers.append(name)
        return col

    tc_col       = headers.index("Test Case ID") + 1
    input_col    = headers.index("Input (EN)") + 1
    expected_col = headers.index("Expected Output") + 1
    criteria_col = headers.index("Validation Criteria") + 1
    actual_col   = get_or_create_col("Actual Output")
    status_col   = get_or_create_col("Status")

    # Append new test cases if not already present
    existing_ids = {ws.cell(row=r, column=tc_col).value for r in range(2, ws.max_row + 1)}
    for case in NEW_CASES:
        if case[0] not in existing_ids:
            row_num = ws.max_row + 1
            for col, val in enumerate(case, start=1):
                ws.cell(row=row_num, column=col, value=val)

    print(f"{'TC ID':<12} {'Status':<6} {'Input':<45} {'Actual Output'}")
    print("-" * 135)

    pass_count = fail_count = 0

    for row_num in range(2, ws.max_row + 1):
        tc_id      = ws.cell(row=row_num, column=tc_col).value
        input_text = ws.cell(row=row_num, column=input_col).value
        expected   = ws.cell(row=row_num, column=expected_col).value
        criteria   = ws.cell(row=row_num, column=criteria_col).value

        if not input_text:
            continue

        result = call_api(str(input_text))
        actual = extract_output(result)
        status = validate(str(input_text), actual, expected, criteria)

        ws.cell(row=row_num, column=actual_col, value=actual)
        ws.cell(row=row_num, column=status_col, value=status)

        if status == "PASS":
            pass_count += 1
        else:
            fail_count += 1

        print(f"{str(tc_id):<12} {status:<6} {str(input_text)[:43]:<45} {actual[:60]}")
        time.sleep(0.3)

    wb.save(OUTPUT_EXCEL)
    print(f"\nDone! PASS: {pass_count} | FAIL: {fail_count} | Total: {pass_count + fail_count}")
    print(f"Saved to: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()


def test_sarvam_full_run():
    main()
