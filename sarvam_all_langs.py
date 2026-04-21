import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import re
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import time

INPUT_EXCEL  = "D:/download/Final_With_New_Scenarios.xlsx"
OUTPUT_EXCEL = "D:/download/Sarvam_All_Languages_Results.xlsx"

API_URL  = "https://api.sarvam.ai/translate"
API_KEY  = "sk_6q9mn91f_NDW17uFDqDE5C5JLDtpR9iyL"
SRC_LANG = "en-IN"
MODEL    = "sarvam-translate:v1"

TARGET_LANGUAGES = [
    ("as-IN",  "Assamese"),
    ("bn-IN",  "Bengali"),
    ("brx-IN", "Bodo"),
    ("hi-IN",  "Hindi"),
    ("en-IN",  "English"),
    ("doi-IN", "Dogri"),
    ("gu-IN",  "Gujarati"),
    ("kn-IN",  "Kannada"),
    ("ks-IN",  "Kashmiri"),
    ("kok-IN", "Konkani"),
    ("mai-IN", "Maithili"),
    ("ml-IN",  "Malayalam"),
    ("mni-IN", "Meitei"),
    ("mr-IN",  "Marathi"),
    ("ne-IN",  "Nepali"),
    ("od-IN",  "Odia"),
    ("pa-IN",  "Punjabi"),
    ("sa-IN",  "Sanskrit"),
    ("sat-IN", "Santali"),
    ("sd-IN",  "Sindhi"),
    ("ta-IN",  "Tamil"),
    ("te-IN",  "Telugu"),
    ("ur-IN",  "Urdu"),
]

NEW_CASES = [
    ('TC_ADD_01', 'Question mark preserved',        'Are you ready?',                                          'Translated with one ?',                        "? not removed or duplicated"),
    ('TC_ADD_02', 'Exclamation preserved',          'Great job!',                                              'Translated with one !',                        'Exclamation mark intact'),
    ('TC_ADD_03', 'Comma preservation',             'Hello, how are you?',                                     'Translated with comma and ?',                  'Comma and ? must match input'),
    ('TC_ADD_04', 'Ellipsis preservation',          'Wait... let me check',                                    'Translated with ellipsis ...',                 'Three dots must remain unchanged'),
    ('TC_ADD_05', 'Semicolon preservation',         'Check one; check two',                                    'Translated with semicolon ;',                  'Semicolon not removed'),
    ('TC_ADD_06', 'Colon preservation',             'Note: This is important',                                 'Translated with colon :',                      'Colon not removed'),
    ('TC_ADD_07', 'Mixed punctuation',              'Hello! How are you? Fine.',                               'Translated with ! ? and . preserved',          '! ? . each count must match input'),
    ('TC_ADD_08', 'Parentheses preservation',       'Result (correct) has been verified',                      'Translated with ( ) intact',                   'Round brackets not removed'),
    ('TC_ADD_09', 'Square brackets',               '[Important] Please review this',                           'Translated with [ ] intact',                   'Square brackets preserved'),
    ('TC_ADD_10', 'Decimal number',                 'Price is 99.99',                                          'Decimal number unchanged',                     'Numeric value 99.99 preserved exactly'),
    ('TC_ADD_11', 'Phone number',                   'Call at 9876543210',                                      '10-digit phone number unchanged',              'Phone number not translated or broken'),
    ('TC_ADD_12', 'Date format DD/MM/YYYY',         'Meeting on 25/04/2024',                                   'Date format unchanged',                        'Date string 25/04/2024 preserved'),
    ('TC_ADD_13', 'Time format AM/PM',              'Meeting at 10:30 AM',                                     'Time format unchanged',                        'Time and AM preserved'),
    ('TC_ADD_14', 'Percentage symbol',              'Score is 95%',                                            'Percentage intact',                            '% symbol and number unchanged'),
    ('TC_ADD_15', 'Negative number',                'Temperature is -5 degrees',                               'Negative sign preserved',                      'Minus sign and number unchanged'),
    ('TC_ADD_16', 'Hashtag handling',               '#TeamWork is the key',                                    'Hashtag label preserved',                      '#TeamWork not translated or broken'),
    ('TC_ADD_17', 'At-mention handling',            '@John please review this',                                '@John preserved as-is',                        '@John not altered'),
    ('TC_ADD_18', 'Dollar currency',                'Price is $500',                                           '$ symbol and number unchanged',                '$ and 500 intact'),
    ('TC_ADD_19', 'Euro currency text',             'Price is 200 euros',                                      'Currency context preserved',                   'Numeric value unchanged'),
    ('TC_ADD_20', 'Asterisk emphasis',              'Note: *important* point',                                 '* symbols preserved',                          'Asterisks not removed'),
    ('TC_ADD_21', 'Forward slash',                  'A/B testing is done',                                     'Slash preserved',                              '/ not removed'),
    ('TC_ADD_22', 'Underscore in identifier',       'file_name is ready',                                      'Underscore preserved',                         'file_name not broken'),
    ('TC_ADD_23', 'Ampersand symbol',               'Terms & Conditions apply',                                '& preserved or spelled out correctly',         '& symbol intact'),
    ('TC_ADD_24', 'Single word input',              'Hello',                                                   'Single word translated',                       'Non-empty output, no extra punctuation'),
    ('TC_ADD_25', 'All caps input',                 'THIS IS IMPORTANT',                                       'Translated meaningfully',                      'Non-empty meaningful output'),
    ('TC_ADD_26', 'Very long sentence',             'The development team has successfully completed the implementation of the new authentication module and all test cases are now passing.',
                                                                                                                'Full translation without truncation',          'Output non-empty and complete'),
    ('TC_ADD_27', 'Repeated words',                 'Very very very good',                                     'Repetition conveyed',                          'Non-empty output with context preserved'),
    ('TC_ADD_28', 'Only numbers input',             '12345',                                                   'Numbers unchanged or numeral form',            'Numeric meaning preserved'),
    ('TC_ADD_29', 'Error code in text',             'Error 404: Page not found',                               '404 preserved, rest translated',               'Numeric code 404 unchanged'),
    ('TC_ADD_30', 'Version number in text',         'Version 2.0.1 has been released',                         'Version number unchanged',                     '2.0.1 intact in output'),
    ('TC_ADD_31', 'Order ID in text',               'Order #ORD-2024-001 is confirmed',                        'Order ID preserved',                           '#ORD-2024-001 unchanged'),
    ('TC_ADD_32', 'Hinglish with question mark',    'Kya yaar this is not working?',                           'Meaning preserved with ?',                     '? preserved, context intact'),
    ('TC_ADD_33', 'Hinglish office context',        'Kal meeting hai office mein',                             'Hindi-English mix translated meaningfully',    'Meaningful Hindi output'),
    ('TC_ADD_34', 'Hinglish request',               'Please dekho this file',                                  'Context preserved in translation',             'Non-empty output'),
    ('TC_ADD_35', 'Exclamation at end',             'Great work done!',                                        '! preserved at end',                           '! intact'),
    ('TC_ADD_36', 'Full stop and exclamation',      'Work is done. Well done!',                                '. and ! both preserved',                       '. and ! count must match input'),
    ('TC_ADD_37', 'No punctuation long',            'This is a simple sentence without any punctuation',       'Translated without adding punctuation',        'No . or ! added to output'),
    ('TC_ADD_38', 'HTML with number',               '<p>Total items: 50</p>',                                  'HTML tags and number preserved',               'Tags intact, 50 unchanged'),
    ('TC_ADD_39', 'Template with OTP',              'Hello {{name}}, your OTP is 1234.',                       'Template and number preserved',                '{{name}} and 1234 unchanged'),
    ('TC_ADD_40', 'URL with query params',          'Visit https://site.com?id=1&page=2',                      'Full URL with params preserved',               'Complete URL unchanged'),
    ('TC_ADD_41', 'Leading spaces',                 '  Hello team',                                            'Leading space handled gracefully',             'Output non-empty and meaningful'),
    ('TC_ADD_42', 'Trailing spaces',                'Hello team  ',                                            'Trailing space handled gracefully',            'Output non-empty and meaningful'),
    ('TC_ADD_43', 'Formal request',                 'Kindly share the updated document at the earliest.',      'Polite tone preserved',                        'Non-empty formal Hindi output'),
    ('TC_ADD_44', 'Question with please',           'Can you please check this issue?',                        'Polite ? preserved',                           '? intact, non-empty output'),
    ('TC_ADD_45', 'Ordinal numbers',                '1st and 2nd tasks are done',                              'Ordinals preserved',                           '1st and 2nd unchanged'),
    ('TC_ADD_46', 'Spelled-out number',             'There are twenty five items',                             'Number context translated',                    'Non-empty output'),
    ('TC_ADD_47', 'Mixed numeral types',            'Call 1800-123-456 or visit Room 101',                     'Phone and room number preserved',              'Numerics unchanged'),
    ('TC_ADD_48', 'Multiple currencies',            'Pay $100 or 8000 rupees',                                 '$ and rupee value preserved',                  'Both numeric values unchanged'),
]

# ── Styles ────────────────────────────────────────────────────────────────────
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
HDR_FILL  = PatternFill("solid", fgColor="4472C4")
HDR_FONT  = Font(bold=True, color="FFFFFF")


# ── API ───────────────────────────────────────────────────────────────────────
def call_api(text, tgt_lang):
    try:
        r = requests.post(
            API_URL,
            headers={"Content-Type": "application/json", "api-subscription-key": API_KEY},
            json={"input": text, "source_language_code": SRC_LANG,
                  "target_language_code": tgt_lang, "model": MODEL},
            timeout=30
        )
        r.raise_for_status()
        data = r.json()
        return data.get("translated_text", str(data))
    except requests.exceptions.RequestException as e:
        return f"ERROR: {e}"


# ── Validation ────────────────────────────────────────────────────────────────
def dot_count(text):
    return text.count(".") + text.count("।") + text.count("।")

def validate(inp, actual, expected_output, criteria):
    if not actual or actual.startswith("ERROR:") or not actual.strip():
        return "FAIL"

    exp  = str(expected_output).lower() if expected_output else ""
    crit = str(criteria).lower()        if criteria        else ""

    # Full stop rules
    if "without full stop" in exp or ("no" in crit and "added" in crit and "." in crit):
        return "PASS" if dot_count(actual) == dot_count(inp) else "FAIL"
    if "one full stop" in exp or "one '.'" in crit:
        return "PASS" if dot_count(actual) == 1 else "FAIL"
    if ("and !" in exp and "." in exp and "both preserved" in exp) or "each count must match" in crit:
        return "PASS" if (dot_count(actual) == dot_count(inp) and
                          actual.count("!") == inp.count("!") and
                          actual.count("?") == inp.count("?")) else "FAIL"
    if "full stop" in exp or "full stop" in crit:
        return "PASS" if dot_count(actual) > 0 else "FAIL"

    # Punctuation
    if "?" in exp or "?" in crit:
        return "PASS" if actual.count("?") == inp.count("?") else "FAIL"
    if "!" in exp or "exclamation" in crit:
        return "PASS" if actual.count("!") == inp.count("!") else "FAIL"
    if ";" in exp or "semicolon" in crit:
        return "PASS" if ";" in actual else "FAIL"
    if "colon" in crit and "not removed" in crit:
        return "PASS" if ":" in actual else "FAIL"
    if "( )" in exp or "round brackets" in crit:
        return "PASS" if ("(" in actual and ")" in actual) else "FAIL"
    if "[ ]" in exp or "square brackets" in crit:
        return "PASS" if ("[" in actual and "]" in actual) else "FAIL"
    if "..." in exp or "ellipsis" in crit:
        return "PASS" if "..." in actual else "FAIL"
    if "%" in exp or "%" in crit:
        return "PASS" if "%" in actual else "FAIL"

    # Token preservation
    tokens = []
    tokens += re.findall(r'https?://\S+', inp)
    tokens += re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', inp)
    tokens += re.findall(r'\{\{[^}]+\}\}', inp)
    tokens += re.findall(r'#[A-Z][A-Z0-9\-]+', inp)
    if any(w in exp or w in crit for w in ("unchanged", "preserved", "intact", "not broken", "not altered")):
        tokens += re.findall(r'\b\d{7,}\b', inp)
        tokens += re.findall(r'\b\d{1,2}/\d{1,2}/\d{4}\b', inp)
        tokens += re.findall(r'\b\d+\.\d+\.\d+\b', inp)
        tokens += re.findall(r'\b\d+\.\d+\b', inp)
        tokens += re.findall(r'\b\d{3,4}\b', inp)
        tokens += re.findall(r'\b\d+(?:st|nd|rd|th)\b', inp)
        tokens += re.findall(r'@\w+', inp)
    if tokens:
        return "PASS" if all(t in actual for t in tokens) else "FAIL"

    # Emoji
    if "emoji" in exp or "emoji" in crit:
        return "PASS" if ([c for c in inp if ord(c) > 0x1F300] ==
                          [c for c in actual if ord(c) > 0x1F300]) else "FAIL"

    if "without adding punctuation" in exp or ("no" in crit and "added" in crit):
        return "PASS" if dot_count(actual) == dot_count(inp) else "FAIL"

    return "PASS" if actual.strip() else "FAIL"


# ── Build test case list from Excel ──────────────────────────────────────────
def load_test_cases(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    headers = [cell.value for cell in ws[1]]
    tc_col   = headers.index("Test Case ID") + 1
    inp_col  = headers.index("Input (EN)") + 1
    exp_col  = headers.index("Expected Output") + 1
    crit_col = headers.index("Validation Criteria") + 1
    scen_col = headers.index("Scenario") + 1

    existing_ids = {ws.cell(row=r, column=tc_col).value for r in range(2, ws.max_row + 1)}
    for case in NEW_CASES:
        if case[0] not in existing_ids:
            row = ws.max_row + 1
            for c, v in enumerate(case, 1):
                ws.cell(row=row, column=c, value=v)

    test_cases = []
    for r in range(2, ws.max_row + 1):
        tc   = ws.cell(row=r, column=tc_col).value
        inp  = ws.cell(row=r, column=inp_col).value
        exp  = ws.cell(row=r, column=exp_col).value
        crit = ws.cell(row=r, column=crit_col).value
        scen = ws.cell(row=r, column=scen_col).value
        if tc and inp:
            test_cases.append((tc, scen, inp, exp, crit))
    return test_cases


# ── Style helpers ─────────────────────────────────────────────────────────────
def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    test_cases = load_test_cases(INPUT_EXCEL)
    total_tc   = len(test_cases)
    total_langs = len(TARGET_LANGUAGES)
    total_calls = total_tc * total_langs

    print(f"Test cases: {total_tc} | Languages: {total_langs} | Total API calls: {total_calls}")
    print("=" * 80)

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)  # remove default sheet

    summary_data = []  # [(lang_code, lang_name, pass, fail)]

    for lang_code, lang_name in TARGET_LANGUAGES:
        print(f"\n── {lang_name} ({lang_code}) ──")
        ws = out_wb.create_sheet(title=f"{lang_name}")

        headers = ["TC ID", "Scenario", "Input (EN)", "Expected Output",
                   "Validation Criteria", f"Actual Output ({lang_name})", "Status"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header_row(ws, len(headers))
        set_col_widths(ws, [12, 28, 38, 38, 35, 45, 8])

        pass_count = fail_count = 0

        for row_idx, (tc_id, scenario, inp, exp, crit) in enumerate(test_cases, start=2):
            actual = call_api(str(inp), lang_code)
            status = validate(str(inp), actual, exp, crit)

            ws.cell(row=row_idx, column=1, value=tc_id)
            ws.cell(row=row_idx, column=2, value=scenario)
            ws.cell(row=row_idx, column=3, value=inp)
            ws.cell(row=row_idx, column=4, value=exp)
            ws.cell(row=row_idx, column=5, value=crit)
            ws.cell(row=row_idx, column=6, value=actual)
            status_cell = ws.cell(row=row_idx, column=7, value=status)
            status_cell.fill = PASS_FILL if status == "PASS" else FAIL_FILL
            status_cell.font = Font(bold=True)
            status_cell.alignment = Alignment(horizontal="center")

            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

            if status == "PASS":
                pass_count += 1
            else:
                fail_count += 1

            print(f"  {str(tc_id):<12} {status:<5} {str(inp)[:40]:<42} {actual[:50]}")
            time.sleep(0.3)

        summary_data.append((lang_code, lang_name, pass_count, fail_count))
        print(f"  >> {lang_name}: PASS={pass_count}  FAIL={fail_count}")

    # ── Summary sheet ──────────────────────────────────────────────────────
    sw = out_wb.create_sheet(title="Summary", index=0)
    sw.append(["Language Code", "Language", "Total", "PASS", "FAIL", "Pass %"])
    style_header_row(sw, 6)
    set_col_widths(sw, [14, 14, 8, 8, 8, 10])

    for lang_code, lang_name, p, f in summary_data:
        total = p + f
        pct   = round(p / total * 100, 1) if total else 0
        row   = sw.max_row + 1
        sw.append([lang_code, lang_name, total, p, f, f"{pct}%"])
        sw.cell(row=row, column=4).fill = PASS_FILL
        sw.cell(row=row, column=5).fill = FAIL_FILL if f > 0 else PASS_FILL
        sw.cell(row=row, column=6).font = Font(bold=True)

    out_wb.save(OUTPUT_EXCEL)
    print(f"\n{'='*80}")
    print(f"All done! Saved to: {OUTPUT_EXCEL}")
    print(f"\nSummary:")
    for lang_code, lang_name, p, f in summary_data:
        print(f"  {lang_name:<12} ({lang_code:<7})  PASS: {p:>2}  FAIL: {f:>2}  ({round(p/(p+f)*100)}%)")


if __name__ == "__main__":
    main()
