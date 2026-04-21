import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

wb = openpyxl.load_workbook('D:/download/Final_With_New_Scenarios.xlsx')
ws = wb['Sheet1']

new_cases = [
    # Punctuation variety
    ('TC_ADD_01', 'Question mark preserved',         'Are you ready?',                                   'Translated with one ?',                              "? not removed or duplicated"),
    ('TC_ADD_02', 'Exclamation preserved',           'Great job!',                                       'Translated with one !',                              'Exclamation mark intact'),
    ('TC_ADD_03', 'Comma preservation',              'Hello, how are you?',                              'Translated with comma and ?',                        'Comma position and ? must match'),
    ('TC_ADD_04', 'Ellipsis preservation',           'Wait... let me check',                             'Translated with ellipsis ...',                       'Three dots must remain unchanged'),
    ('TC_ADD_05', 'Semicolon preservation',          'Check one; check two',                             'Translated with semicolon ;',                        'Semicolon not removed'),
    ('TC_ADD_06', 'Colon preservation',              'Note: This is important',                          'Translated with colon :',                            'Colon not removed'),
    ('TC_ADD_07', 'Mixed punctuation',               'Hello! How are you? Fine.',                        'Translated with ! ? and . preserved',                '! ? . each count must match input'),
    ('TC_ADD_08', 'Parentheses preservation',        'Result (correct) has been verified',               'Translated with ( ) intact',                         'Round brackets not removed'),
    ('TC_ADD_09', 'Square brackets',                 '[Important] Please review this',                   'Translated with [ ] intact',                         'Square brackets preserved'),

    # Number formats
    ('TC_ADD_10', 'Decimal number',                  'Price is 99.99',                                   'Decimal number unchanged',                           'Numeric value 99.99 preserved exactly'),
    ('TC_ADD_11', 'Phone number',                    'Call at 9876543210',                               '10-digit phone number unchanged',                    'Phone number not translated or broken'),
    ('TC_ADD_12', 'Date format DD/MM/YYYY',          'Meeting on 25/04/2024',                            'Date format unchanged',                              'Date string 25/04/2024 preserved'),
    ('TC_ADD_13', 'Time format AM/PM',               'Meeting at 10:30 AM',                              'Time format unchanged',                              'Time and AM preserved'),
    ('TC_ADD_14', 'Percentage symbol',               'Score is 95%',                                     'Percentage intact',                                  '% symbol and number unchanged'),
    ('TC_ADD_15', 'Negative number',                 'Temperature is -5 degrees',                        'Negative sign preserved',                            'Minus sign and number unchanged'),

    # Special symbols
    ('TC_ADD_16', 'Hashtag handling',                '#TeamWork is the key',                             'Hashtag label preserved',                            '#TeamWork not translated or broken'),
    ('TC_ADD_17', 'At-mention handling',             '@John please review this',                         '@John preserved as-is',                              '@John not altered'),
    ('TC_ADD_18', 'Dollar currency',                 'Price is $500',                                    '$ symbol and number unchanged',                      '$ and 500 intact'),
    ('TC_ADD_19', 'Euro currency',                   'Price is 200 euros',                               'Currency context preserved',                         'Numeric value unchanged'),
    ('TC_ADD_20', 'Asterisk emphasis',               'Note: *important* point',                          '* symbols preserved',                                'Asterisks not removed'),
    ('TC_ADD_21', 'Forward slash',                   'A/B testing is done',                              'Slash preserved',                                    '/ not removed'),
    ('TC_ADD_22', 'Underscore in identifier',        'file_name is ready',                               'Underscore preserved',                               'file_name not broken'),
    ('TC_ADD_23', 'Ampersand symbol',                'Terms & Conditions apply',                         '& preserved or spelled out correctly',               '& symbol intact'),

    # Edge cases
    ('TC_ADD_24', 'Single word input',               'Hello',                                            'Single word translated',                             'Non-empty output, no extra punctuation added'),
    ('TC_ADD_25', 'All caps input',                  'THIS IS IMPORTANT',                                'Translated meaningfully',                            'Non-empty meaningful output'),
    ('TC_ADD_26', 'Very long sentence',              'The development team has successfully completed the implementation of the new authentication module and all test cases are now passing.',
                                                                                                         'Full translation without truncation',                'Output non-empty and complete'),
    ('TC_ADD_27', 'Repeated words',                  'Very very very good',                              'Repetition conveyed',                                'Non-empty output with context preserved'),
    ('TC_ADD_28', 'Only numbers input',              '12345',                                            'Numbers unchanged or numeral form',                  'Numeric meaning preserved'),

    # Code-like content
    ('TC_ADD_29', 'Error code in text',              'Error 404: Page not found',                        '404 preserved, rest translated',                     'Numeric code 404 unchanged'),
    ('TC_ADD_30', 'Version number in text',          'Version 2.0.1 has been released',                  'Version number unchanged',                           '2.0.1 intact in output'),
    ('TC_ADD_31', 'Order ID in text',                'Order #ORD-2024-001 is confirmed',                 'Order ID preserved',                                 '#ORD-2024-001 unchanged'),

    # More Hinglish
    ('TC_ADD_32', 'Hinglish with question mark',     'Kya yaar this is not working?',                   'Meaning preserved with ?',                           '? preserved, context intact'),
    ('TC_ADD_33', 'Hinglish office context',         'Kal meeting hai office mein',                      'Hindi-English mix translated meaningfully',           'Meaningful Hindi output'),
    ('TC_ADD_34', 'Hinglish request',                'Please dekho this file',                           'Context preserved in translation',                   'Non-empty output'),

    # Emoji edge cases
    ('TC_ADD_35', 'Emoji at start of sentence',      'Great work done!',                                 'Trailing ! preserved',                               '! intact'),
    ('TC_ADD_36', 'Emoji at end',                    'Work is done. Well done!',                        '. and ! both preserved',                             '. and ! count must match'),
    ('TC_ADD_37', 'Repeated emojis',                 'Ha ha ha good work',                              'Meaning conveyed',                                   'Non-empty output'),

    # HTML + combos
    ('TC_ADD_38', 'HTML with number',                '<p>Total items: 50</p>',                           'HTML tags and number preserved',                     'Tags intact, 50 unchanged'),
    ('TC_ADD_39', 'Template with OTP',               'Hello {{name}}, your OTP is 1234.',               'Template and number preserved with .',                '{{name}} and 1234 unchanged'),
    ('TC_ADD_40', 'URL with query params',            'Visit https://site.com?id=1&page=2',              'Full URL with params preserved',                     'Complete URL unchanged'),

    # Multiline / whitespace
    ('TC_ADD_41', 'Leading space',                   '  Hello team',                                     'Leading space handled gracefully',                   'Output non-empty, meaningful'),
    ('TC_ADD_42', 'Trailing space',                  'Hello team  ',                                     'Trailing space handled gracefully',                  'Output non-empty, meaningful'),
    ('TC_ADD_43', 'Extra internal spaces',           'Hello   world',                                    'Extra spaces handled gracefully',                    'Output non-empty, meaningful'),

    # Polite / formal tones
    ('TC_ADD_44', 'Formal request',                  'Kindly share the updated document at the earliest.',  'Polite tone preserved',                         'Non-empty, formal Hindi output'),
    ('TC_ADD_45', 'Question with please',            'Can you please check this issue?',                 'Polite ? preserved',                                 '? intact, non-empty output'),

    # Numbers mixed with text
    ('TC_ADD_46', 'Ordinal numbers',                 '1st and 2nd tasks are done',                       'Ordinals preserved',                                 '1st and 2nd unchanged'),
    ('TC_ADD_47', 'Spelled-out number',              'There are twenty five items',                      'Number context translated',                          'Non-empty output'),
    ('TC_ADD_48', 'Mixed numeral types',             'Call 1800-123-456 or visit Room 101',              'Phone and room number preserved',                    'Numerics unchanged'),
]

start_row = ws.max_row + 1

for i, case in enumerate(new_cases):
    row_num = start_row + i
    for col, val in enumerate(case, start=1):
        ws.cell(row=row_num, column=col, value=val)

out = 'D:/download/Final_With_New_Scenarios_Extended.xlsx'
wb.save(out)
print(f'Added {len(new_cases)} new test cases. Total rows now: {ws.max_row - 1}')
print(f'Saved to: {out}')
